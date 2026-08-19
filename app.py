"""
Rahul Fire Data Entry - Full Edition (Revised)
Multi-user Google Sheets + Google Drive app with:
  - Single Main Admin (full control) + Employee accounts (restricted)
  - Employees can edit their OWN entries for up to 7 days; after that,
    only Admin can edit. Employees can never edit others' entries.
  - Up to 2 file attachments (JPEG/JPG/PDF) per Purchase entry and per
    unique Challan, stored in Google Drive under:
      SITE_{site_name}/Purchase  or  SITE_{site_name}/Challan
  - File names are updated when entries are edited.
  - On deletion, files are renamed with an EXTRA_ prefix instead of being deleted.
  - Mobile-friendly UI and role-restricted navigation.
  - Employees cannot export Excel.

SHEET SCHEMA (must match exactly):
  Purchases:
    id | entry_date | purchaser | vendor | invoice_no | amount | payment_mode |
    payment_detail | payment_date | site_name | challan_number | notes |
    created_by | created_at | file1_link | file2_link

  Challans:
    id | challan_number | challan_date | site_name | vehicle_number | driver_name |
    created_by | created_at | file1_link | file2_link

  ChallanItems:
    id | challan_id | sr_no | description | qty_nos | qty_meters | billable | created_by

  Users:
    username | password_hash | full_name | role

ENVIRONMENT VARIABLES:
  GOOGLE_CREDENTIALS_JSON   (required)
  APP_USERNAME              (Main Admin username)
  APP_PASSWORD_HASH         (Main Admin password hash)
  GOOGLE_DRIVE_FOLDER_NAME  (optional, default: "Rahul Fire Data Entry - Attachments")
  GOOGLE_SHEET_NAME         (optional, default: "Rahul Fire Data Entry - Database")

requirements.txt must include:
  flask
  gunicorn
  openpyxl
  gspread
  google-auth
  google-api-python-client
"""

import os
import io
import json
import hashlib
import secrets
import threading
import re as _re
from functools import wraps
from datetime import datetime, timedelta, timezone
from flask import Flask, request, redirect, url_for, send_file, session, render_template_string
from markupsafe import escape

APP_TITLE = "Rahul Fire Data Entry"
EDIT_WINDOW_DAYS = 7
MAX_SITES = 20
MAX_EMPLOYEES = 20
MAX_FILES_PER_ENTRY = 2
ALLOWED_EXTENSIONS = {"jpg", "jpeg", "pdf"}
PAYMENT_MODES = ["Cash", "Bank Transfer", "UPI", "Cheque", "Credit Card", "Debit Card"]
BILLABLE_OPTIONS = ["Billable", "Non-Billable"]

PURCHASE_HEADERS = [
    "id", "entry_date", "purchaser", "vendor", "invoice_no", "amount",
    "payment_mode", "payment_detail", "payment_date", "site_name",
    "challan_number", "notes", "created_by", "created_at", "file1_link", "file2_link"
]
CHALLAN_HEADERS = [
    "id", "challan_number", "challan_date", "site_name",
    "vehicle_number", "driver_name", "created_by", "created_at",
    "file1_link", "file2_link"
]
ITEM_HEADERS = [
    "id", "challan_id", "sr_no", "description",
    "qty_nos", "qty_meters", "billable", "created_by"
]
USER_HEADERS = ["username", "password_hash", "full_name", "role"]

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", secrets.token_hex(32))
app.permanent_session_lifetime = timedelta(days=30)
app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024

WRITE_LOCK = threading.Lock()


def hash_password(password, salt=None):
    if salt is None:
        salt = secrets.token_hex(16)
    h = hashlib.sha256((salt + password).encode()).hexdigest()
    return f"{salt}${h}"


def verify_password(password, stored):
    try:
        salt, h = stored.split("$")
        return hashlib.sha256((salt + password).encode()).hexdigest() == h
    except Exception:
        return False

  # ---------------------------------------------------------------------------
# Google Sheets & Drive connections
# ---------------------------------------------------------------------------
_gs_client = None
_sheet = None
_drive_service = None
_drive_folder_id = None


def get_sheet():
    global _gs_client, _sheet
    if _sheet is not None:
        return _sheet
    import gspread
    from google.oauth2.service_account import Credentials
    creds_json = os.environ.get("GOOGLE_CREDENTIALS_JSON")
    sheet_name = os.environ.get("GOOGLE_SHEET_NAME", "Rahul Fire Data Entry - Database")
    if not creds_json:
        raise RuntimeError("GOOGLE_CREDENTIALS_JSON environment variable is not set.")
    creds_dict = json.loads(creds_json)
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    _gs_client = gspread.authorize(creds)
    _sheet = _gs_client.open(sheet_name)
    return _sheet


def get_drive_service():
    global _drive_service
    if _drive_service is not None:
        return _drive_service
    from googleapiclient.discovery import build
    from google.oauth2.service_account import Credentials
    creds_json = os.environ.get("GOOGLE_CREDENTIALS_JSON")
    creds_dict = json.loads(creds_json)
    scopes = ["https://www.googleapis.com/auth/drive"]
    creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    _drive_service = build("drive", "v3", credentials=creds)
    return _drive_service


def get_drive_folder_id():
    global _drive_folder_id
    if _drive_folder_id:
        return _drive_folder_id
    service = get_drive_service()
    folder_name = os.environ.get(
        "GOOGLE_DRIVE_FOLDER_NAME",
        "Rahul Fire Data Entry - Attachments"
    )
    query = f"name='{folder_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    results = service.files().list(q=query, fields="files(id, name)").execute()
    files = results.get("files", [])
    if files:
        _drive_folder_id = files[0]["id"]
    else:
        metadata = {"name": folder_name, "mimeType": "application/vnd.google-apps.folder"}
        folder = service.files().create(body=metadata, fields="id").execute()
        _drive_folder_id = folder["id"]
    return _drive_folder_id


def get_site_folder_id(site_name):
    """
    Get or create a Drive folder named 'SITE_{site_name}'.
    Returns the folder ID.
    """
    service = get_drive_service()
    folder_name = f"SITE_{site_name}"
    query = f"name='{folder_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    results = service.files().list(q=query, fields="files(id, name)").execute()
    files = results.get("files", [])
    if files:
        return files[0]["id"]
    parent_id = get_drive_folder_id()
    metadata = {
        "name": folder_name,
        "mimeType": "application/vnd.google-apps.folder",
        "parents": [parent_id]
    }
    folder = service.files().create(body=metadata, fields="id").execute()
    return folder["id"]


def get_or_create_subfolder(parent_id, name):
    """
    Get or create a subfolder with given name under parent_id.
    Returns the subfolder ID.
    """
    service = get_drive_service()
    query = f"name='{name}' and mimeType='application/vnd.google-apps.folder' and trashed=false and '{parent_id}' in parents"
    results = service.files().list(q=query, fields="files(id)").execute()
    files = results.get("files", [])
    if files:
        return files[0]["id"]
    metadata = {
        "name": name,
        "mimeType": "application/vnd.google-apps.folder",
        "parents": [parent_id]
    }
    folder = service.files().create(body=metadata, fields="id").execute()
    return folder["id"]


def rename_drive_file(file_id, new_name):
    """
    Rename an existing Drive file to new_name.
    """
    service = get_drive_service()
    service.files().update(fileId=file_id, body={"name": new_name}).execute()


def extract_drive_file_id(file_url):
    """
    Extract Drive file ID from a URL like:
    https://drive.google.com/file/d/FILE_ID/view
    Returns FILE_ID or None if not found.
    """
    if not file_url:
        return None
    try:
        parts = file_url.split("/d/")
        if len(parts) < 2:
            return None
        file_id_part = parts[1].split("/")[0]
        return file_id_part
    except Exception:
        return None


def upload_file_to_drive_with_site(file_obj, filename, mimetype, site_name, entry_type):
    """
    Upload a file to Drive under:
      SITE_{site_name}/Purchase  or  SITE_{site_name}/Challan
    entry_type: "PUR" or "CH"
    If a file with the same name exists in that folder, overwrite it.
    Returns the public viewer URL.
    """
    service = get_drive_service()
    site_folder_id = get_site_folder_id(site_name)
    subfolder_name = "Purchase" if entry_type == "PUR" else "Challan"
    subfolder_id = get_or_create_subfolder(site_folder_id, subfolder_name)

    query = f"name='{filename}' and trashed=false and '{subfolder_id}' in parents"
    results = service.files().list(q=query, fields="files(id)").execute()
    existing = results.get("files", [])

    file_bytes = file_obj.read()
    from googleapiclient.http import MediaIoBaseUpload
    media = MediaIoBaseUpload(io.BytesIO(file_bytes), mimetype=mimetype, resumable=False)

    if existing:
        file_id = existing[0]["id"]
        service.files().update(fileId=file_id, media_body=media).execute()
    else:
        metadata = {"name": filename, "parents": [subfolder_id]}
        created = service.files().create(body=metadata, media_body=media, fields="id").execute()
        file_id = created["id"]
        service.permissions().create(
            fileId=file_id,
            body={"type": "anyone", "role": "reader"}
        ).execute()

    return f"https://drive.google.com/file/d/{file_id}/view"


def validate_uploaded_files(file_list):
    real_files = [f for f in file_list if f and f.filename]
    if len(real_files) > MAX_FILES_PER_ENTRY:
        return False, f"You can upload a maximum of {MAX_FILES_PER_ENTRY} files."
    for f in real_files:
        ext = f.filename.rsplit(".", 1)[-1].lower() if "." in f.filename else ""
        if ext not in ALLOWED_EXTENSIONS:
            return False, f"'{f.filename}' is not allowed. Only JPG, JPEG, and PDF files can be uploaded."
    return True, "OK"


def generate_attachment_filename(entry_type, entry_id, slot_number, original_filename):
    ext = original_filename.rsplit(".", 1)[-1].lower() if "." in original_filename else "bin"
    safe_entry_id = "".join(c for c in str(entry_id) if c.isalnum() or c in "-_")
    if entry_type == "CH" and safe_entry_id.upper().startswith("CH-"):
        safe_entry_id = safe_entry_id[3:]
    return f"{entry_type}-{safe_entry_id}-slot{slot_number}.{ext}"


def process_file_uploads_with_site(files, entry_type, entry_id, site_name):
    """
    files: request.files.getlist('attachments')
    entry_type: "PUR" or "CH"
    entry_id: purchase ID or challan number (string)
    site_name: site name from the entry
    Returns (file1_link, file2_link, error)
    """
    real_files = [f for f in files if f and f.filename]
    ok, err = validate_uploaded_files(real_files)
    if not ok:
        return None, None, err

    links = [None, None]
    for i, f in enumerate(real_files[:MAX_FILES_PER_ENTRY]):
        slot = i + 1
        fname = generate_attachment_filename(entry_type, entry_id, slot, f.filename)
        ext = fname.rsplit(".", 1)[-1]
        mimetype = "application/pdf" if ext == "pdf" else "image/jpeg"
        try:
            link = upload_file_to_drive_with_site(f, fname, mimetype, site_name, entry_type)
            links[i] = link
        except Exception as e:
            return None, None, f"File upload failed: {e}"
    return links[0], links[1], None


def rename_files_for_purchase_edit(row_id, old_row, new_row):
    """
    When a purchase entry is edited, rename associated Drive files to match the new entry ID.
    """
    for slot, old_link in [(1, old_row.get("file1_link")), (2, old_row.get("file2_link"))]:
        if not old_link:
            continue
        file_id = extract_drive_file_id(old_link)
        if not file_id:
            continue
        old_name_parts = old_link.split("/")[-1].split(".")
        ext = old_name_parts[-1] if len(old_name_parts) > 1 else "bin"
        new_name = f"PUR-{row_id}-slot{slot}.{ext}"
        rename_drive_file(file_id, new_name)


def rename_files_for_challan_edit(challan_number, old_row, new_row):
    """
    When a challan header is edited, rename associated Drive files to match the (possibly) new challan number.
    """
    for slot, old_link in [(1, old_row.get("file1_link")), (2, old_row.get("file2_link"))]:
        if not old_link:
            continue
        file_id = extract_drive_file_id(old_link)
        if not file_id:
            continue
        safe_ch = "".join(c for c in str(challan_number) if c.isalnum() or c in "-_")
        old_name_parts = old_link.split("/")[-1].split(".")
        ext = old_name_parts[-1] if len(old_name_parts) > 1 else "bin"
        new_name = f"CH-{safe_ch}-slot{slot}.{ext}"
        rename_drive_file(file_id, new_name)


def rename_files_for_delete(entry_type, old_row):
    """
    When an entry is deleted, rename its Drive files with an 'EXTRA_' prefix instead of deleting them.
    """
    id_value = old_row.get("id") if entry_type == "PUR" else old_row.get("challan_number")
    for slot, old_link in [(1, old_row.get("file1_link")), (2, old_row.get("file2_link"))]:
        if not old_link:
            continue
        file_id = extract_drive_file_id(old_link)
        if not file_id:
            continue
        old_name_parts = old_link.split("/")[-1].split(".")
        ext = old_name_parts[-1] if len(old_name_parts) > 1 else "bin"
        safe_id = "".join(c for c in str(id_value) if c.isalnum() or c in "-_")
        new_name = f"EXTRA-{entry_type}-{safe_id}-slot{slot}.{ext}"
        rename_drive_file(file_id, new_name)

  # ---------------------------------------------------------------------------
# Google Sheets helpers
# ---------------------------------------------------------------------------
def get_ws(tab_name):
    return get_sheet().worksheet(tab_name)


def read_all_records(tab_name):
    return get_ws(tab_name).get_all_records()


def next_id(tab_name):
    records = read_all_records(tab_name)
    ids = [int(r["id"]) for r in records if str(r.get("id", "")).strip().isdigit()]
    return (max(ids) + 1) if ids else 1


def append_row(tab_name, headers, row_dict):
    ws = get_ws(tab_name)
    row = [row_dict.get(h, "") for h in headers]
    ws.append_row(row, value_input_option="USER_ENTERED")


def find_row_index_by_id(tab_name, row_id):
    ws = get_ws(tab_name)
    ids_col = ws.col_values(1)
    for idx, val in enumerate(ids_col, start=1):
        if idx == 1:
            continue
        if str(val).strip() == str(row_id):
            return idx
    return None


def update_row_by_id(tab_name, headers, row_id, row_dict):
    ws = get_ws(tab_name)
    idx = find_row_index_by_id(tab_name, row_id)
    if idx is None:
        raise ValueError(f"Row with id={row_id} not found in {tab_name}")
    row = [row_dict.get(h, "") for h in headers]
    ws.update(f"A{idx}:{chr(64 + len(headers))}{idx}", [row])


def delete_row_by_id(tab_name, row_id):
    ws = get_ws(tab_name)
    idx = find_row_index_by_id(tab_name, row_id)
    if idx is not None:
        ws.delete_rows(idx)


# ---------------------------------------------------------------------------
# Users & Authentication
# ---------------------------------------------------------------------------
def get_all_users():
    try:
        return read_all_records("Users")
    except Exception:
        return []


def verify_user_login(username, password):
    for u in get_all_users():
        if u.get("username") == username and verify_password(password, u.get("password_hash", "")):
            return {
                "username": u["username"],
                "full_name": u.get("full_name", u["username"]),
                "role": u.get("role", "Employee")
            }
    admin_user = os.environ.get("APP_USERNAME", "")
    admin_hash = os.environ.get("APP_PASSWORD_HASH", "")
    if admin_user and username == admin_user and verify_password(password, admin_hash):
        return {"username": admin_user, "full_name": "Main Admin", "role": "Admin"}
    return None


def add_user(username, password, full_name, role="Employee"):
    users = get_all_users()
    if any(u["username"] == username for u in users):
        return False, "That username already exists."
    if len(users) >= 20:
        return False, "Maximum of 20 users reached."
    pw_hash = hash_password(password)
    append_row("Users", USER_HEADERS, {
        "username": username,
        "password_hash": pw_hash,
        "full_name": full_name,
        "role": role
    })
    return True, "User created."


def delete_user(username):
    ws = get_ws("Users")
    usernames = ws.col_values(1)
    for idx, val in enumerate(usernames, start=1):
        if val == username:
            ws.delete_rows(idx)
            return True, f"User '{username}' deleted."
    return False, "User not found."


LOGIN_PAGE = """
<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Login - """ + APP_TITLE + """</title>
<style>
:root {
  --bg: #f5f5f7;
  --card-bg: #ffffff;
  --primary: #7A1F1F;
  --text: #111111;
  --muted: #666666;
}
* { box-sizing: border-box; }
body {
  font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Arial, sans-serif;
  margin: 0;
  background: var(--bg);
  color: var(--text);
  display: flex;
  justify-content: center;
  align-items: center;
  min-height: 100vh;
  padding: 1rem;
}
.box {
  background: var(--card-bg);
  padding: 1.25rem;
  border-radius: 12px;
  box-shadow: 0 2px 8px rgba(0,0,0,0.12);
  width: 100%;
  max-width: 360px;
}
h2 {
  color: var(--primary);
  text-align: center;
  margin: 0 0 0.75rem 0;
  font-size: 1.35rem;
}
label {
  display: block;
  margin-top: 0.6rem;
  margin-bottom: 0.25rem;
  font-weight: 600;
  font-size: 0.9rem;
}
input {
  width: 100%;
  padding: 0.65rem 0.7rem;
  font-size: 1rem;
  border-radius: 8px;
  border: 1px solid #ccc;
  background: #fff;
}
button {
  width: 100%;
  padding: 0.7rem 0.8rem;
  font-size: 1.05rem;
  font-weight: 700;
  border-radius: 8px;
  min-height: 48px;
  margin-top: 0.75rem;
  background: var(--primary);
  color: #fff;
  border: none;
  cursor: pointer;
}
.error {
  color: #b00020;
  text-align: center;
  margin-bottom: 0.6rem;
  font-size: 0.9rem;
}
@media (max-width: 400px) {
  .box {
    padding: 1rem;
  }
  h2 {
    font-size: 1.2rem;
  }
}
</style>
</head>
<body>
<div class="box">
  <h2>""" + APP_TITLE + """</h2>
  {% if error %}<p class="error">{{ error }}</p>{% endif %}
  <form method="post">
    <label for="username">Username</label>
    <input type="text" id="username" name="username" placeholder="Username" required autofocus>
    <label for="password">Password</label>
    <input type="password" id="password" name="password" placeholder="Password" required>
    <button type="submit">Login</button>
  </form>
</div>
</body>
</html>
"""


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        username = request.form.get("username", "")
        password = request.form.get("password", "")
        user = verify_user_login(username, password)
        if user:
            session["logged_in"] = True
            session["username"] = user["username"]
            session["full_name"] = user["full_name"]
            session["role"] = user["role"]
            session.permanent = True
            return redirect(url_for("dashboard"))
        error = "Invalid username or password."
    return render_template_string(LOGIN_PAGE, error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapper


def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        if session.get("role") != "Admin":
            return page("Access Denied", "<p>Only the Main Admin can access this page.</p>")
        return f(*args, **kwargs)
    return wrapper


def current_user():
    return session.get("username", "unknown")


def is_admin():
    return session.get("role") == "Admin"


def is_employee():
    return session.get("role") == "Employee"


def can_modify_entry(row_created_by, row_created_at_iso):
    """Returns (allowed: bool, reason: str or None)."""
    if is_admin():
        return True, None
    if row_created_by != current_user():
        return False, "You can only edit entries you created yourself."
    if not row_created_at_iso:
        return False, "This entry has no creation timestamp on record; only the Main Admin can edit it."
    try:
        created_dt = datetime.fromisoformat(row_created_at_iso)
        if created_dt.tzinfo is None:
            created_dt = created_dt.replace(tzinfo=timezone.utc)
    except Exception:
        return False, "Could not read this entry's creation date; only the Main Admin can edit it."
    age = datetime.now(timezone.utc) - created_dt
    if age > timedelta(days=EDIT_WINDOW_DAYS):
        return False, f"This entry is older than {EDIT_WINDOW_DAYS} days and can now only be edited by the Main Admin."
    return True, None


def now_iso():
    return datetime.now(timezone.utc).isoformat()


# ---------------------------------------------------------------------------
# Sites & Employees
# ---------------------------------------------------------------------------
def get_sites():
    return [r["name"] for r in read_all_records("Sites") if r.get("name")]


def get_employees():
    return [r["name"] for r in read_all_records("Employees") if r.get("name")]


def add_site(name):
    with WRITE_LOCK:
        sites = get_sites()
        if len(sites) >= MAX_SITES:
            return False, f"Maximum of {MAX_SITES} sites reached."
        if name in sites:
            return False, "A site with that name already exists."
        get_ws("Sites").append_row([name])
        return True, "Site added."


def rename_site(old_name, new_name):
    with WRITE_LOCK:
        ws = get_ws("Sites")
        names = ws.col_values(1)
        for idx, val in enumerate(names, start=1):
            if val == old_name:
                ws.update_cell(idx, 1, new_name)
                break
        for tab, headers, col in [("Purchases", PURCHASE_HEADERS, "site_name"), ("Challans", CHALLAN_HEADERS, "site_name")]:
            records = read_all_records(tab)
            ws2 = get_ws(tab)
            col_idx = headers.index(col) + 1
            for i, rec in enumerate(records, start=2):
                if rec.get(col) == old_name:
                    ws2.update_cell(i, col_idx, new_name)


def delete_site(name, force=False):
    with WRITE_LOCK:
        purchases = [r for r in read_all_records("Purchases") if r.get("site_name") == name]
        challans = [r for r in read_all_records("Challans") if r.get("site_name") == name]
        if (purchases or challans) and not force:
            return False, f"'{name}' has {len(purchases)} purchase(s) and {len(challans)} challan(s) linked. Use Force Delete."
        ws = get_ws("Sites")
        names = ws.col_values(1)
        for idx, val in enumerate(names, start=1):
            if val == name:
                ws.delete_rows(idx)
                break
        return True, f"Site '{name}' deleted."


def add_employee(name):
    with WRITE_LOCK:
        emps = get_employees()
        if len(emps) >= MAX_EMPLOYEES:
            return False, f"Maximum of {MAX_EMPLOYEES} employees reached."
        if name in emps:
            return False, "An employee with that name already exists."
        get_ws("Employees").append_row([name])
        return True, "Employee added."


def rename_employee(old_name, new_name):
    with WRITE_LOCK:
        ws = get_ws("Employees")
        names = ws.col_values(1)
        for idx, val in enumerate(names, start=1):
            if val == old_name:
                ws.update_cell(idx, 1, new_name)
                break
        records = read_all_records("Purchases")
        ws2 = get_ws("Purchases")
        col_idx = PURCHASE_HEADERS.index("purchaser") + 1
        for i, rec in enumerate(records, start=2):
            if rec.get("purchaser") == old_name:
                ws2.update_cell(i, col_idx, new_name)


def delete_employee(name):
    with WRITE_LOCK:
        ws = get_ws("Employees")
        names = ws.col_values(1)
        for idx, val in enumerate(names, start=1):
            if val == name:
                ws.delete_rows(idx)
                break
        return True, f"Employee '{name}' deleted."

  # ---------------------------------------------------------------------------
# Purchases
# ---------------------------------------------------------------------------
def get_purchases(site_name=None):
    records = read_all_records("Purchases")
    if site_name and site_name != "ALL SITES":
        records = [r for r in records if r.get("site_name") == site_name]
    records.sort(key=lambda r: (str(r.get("entry_date", "")), int(r.get("id", 0)) if str(r.get("id", "")).isdigit() else 0))
    return records


def check_duplicate_purchase(entry_date, vendor, invoice_no, amount, exclude_id=None):
    for r in read_all_records("Purchases"):
        if (str(r.get("entry_date")) == entry_date and r.get("vendor") == vendor and
            r.get("invoice_no") == invoice_no and float(r.get("amount") or 0) == float(amount)):
            if exclude_id is not None and str(r.get("id")) == str(exclude_id):
                continue
            return True
    return False


def insert_purchase(record, created_by, files=None):
    with WRITE_LOCK:
        row_id = next_id("Purchases")
        record["id"] = row_id
        record["created_by"] = created_by
        record["created_at"] = now_iso()
        record.setdefault("file1_link", "")
        record.setdefault("file2_link", "")

        if files:
            real_files = [f for f in files if f and f.filename]
            if real_files:
                f1, f2, uerr = process_file_uploads_with_site(
                    real_files, "PUR", str(row_id), record["site_name"]
                )
                if uerr:
                    raise ValueError(uerr)
                if f1:
                    record["file1_link"] = f1
                if f2:
                    record["file2_link"] = f2

        append_row("Purchases", PURCHASE_HEADERS, record)
        return row_id


def update_purchase(row_id, record, files=None):
    with WRITE_LOCK:
        existing = get_purchase_row(row_id)
        record["id"] = row_id
        record["created_by"] = existing.get("created_by", "") if existing else ""
        record["created_at"] = existing.get("created_at", "") if existing else ""

        if files:
            real_files = [f for f in files if f and f.filename]
            if real_files:
                f1, f2, uerr = process_file_uploads_with_site(
                    real_files, "PUR", str(row_id), record["site_name"]
                )
                if uerr:
                    raise ValueError(uerr)
                if f1:
                    record["file1_link"] = f1
                if f2:
                    record["file2_link"] = f2
        else:
            record["file1_link"] = record.get("file1_link") or (existing.get("file1_link", "") if existing else "")
            record["file2_link"] = record.get("file2_link") or (existing.get("file2_link", "") if existing else "")

        rename_files_for_purchase_edit(row_id, existing or {}, record)
        update_row_by_id("Purchases", PURCHASE_HEADERS, row_id, record)


def delete_purchase(row_id):
    with WRITE_LOCK:
        existing = get_purchase_row(row_id)
        if existing:
            rename_files_for_delete("PUR", existing)
        delete_row_by_id("Purchases", row_id)


def get_purchase_row(row_id):
    for r in read_all_records("Purchases"):
        if str(r.get("id")) == str(row_id):
            return r
    return None


# ---------------------------------------------------------------------------
# Challans
# ---------------------------------------------------------------------------
def _numeric_key(challan_no):
    m = _re.search(r"\d+", str(challan_no or ""))
    return int(m.group()) if m else 0


def challan_date_mismatch(challan_number, challan_date, exclude_id=None):
    for r in read_all_records("Challans"):
        if r.get("challan_number") == challan_number:
            if exclude_id is not None and str(r.get("id")) == str(exclude_id):
                continue
            if str(r.get("challan_date")) != challan_date:
                return r.get("challan_date")
    return None


def get_or_create_challan(challan_number, challan_date, site_name, vehicle_number, driver_name, created_by):
    with WRITE_LOCK:
        records = read_all_records("Challans")
        for r in records:
            if r.get("challan_number") == challan_number:
                update_row_by_id("Challans", CHALLAN_HEADERS, r["id"], {
                    "id": r["id"],
                    "challan_number": challan_number,
                    "challan_date": challan_date,
                    "site_name": site_name,
                    "vehicle_number": vehicle_number,
                    "driver_name": driver_name,
                    "created_by": r.get("created_by", created_by),
                    "created_at": r.get("created_at", now_iso()),
                    "file1_link": r.get("file1_link", ""),
                    "file2_link": r.get("file2_link", "")
                })
                return r["id"]
        row_id = next_id("Challans")
        append_row("Challans", CHALLAN_HEADERS, {
            "id": row_id,
            "challan_number": challan_number,
            "challan_date": challan_date,
            "site_name": site_name,
            "vehicle_number": vehicle_number,
            "driver_name": driver_name,
            "created_by": created_by,
            "created_at": now_iso(),
            "file1_link": "",
            "file2_link": ""
        })
        return row_id


def update_challan_header(challan_id, challan_date, site_name, vehicle_number, driver_name, file1_link=None, file2_link=None, files=None):
    with WRITE_LOCK:
        challan = get_challan(challan_id)
        if not challan:
            raise ValueError(f"Challan {challan_id} not found")

        if files:
            real_files = [f for f in files if f and f.filename]
            if real_files:
                f1, f2, uerr = process_file_uploads_with_site(
                    real_files, "CH", str(challan["challan_number"]), site_name
                )
                if uerr:
                    raise ValueError(uerr)
                if f1:
                    challan["file1_link"] = f1
                if f2:
                    challan["file2_link"] = f2
        else:
            if file1_link:
                challan["file1_link"] = file1_link
            if file2_link:
                challan["file2_link"] = file2_link

        rename_files_for_challan_edit(challan["challan_number"], challan, challan)

        update_row_by_id("Challans", CHALLAN_HEADERS, challan_id, {
            "id": challan_id,
            "challan_number": challan["challan_number"],
            "challan_date": challan_date,
            "site_name": site_name,
            "vehicle_number": vehicle_number,
            "driver_name": driver_name,
            "created_by": challan.get("created_by", ""),
            "created_at": challan.get("created_at", ""),
            "file1_link": challan.get("file1_link", ""),
            "file2_link": challan.get("file2_link", "")
        })


def get_challan(challan_id):
    for r in read_all_records("Challans"):
        if str(r.get("id")) == str(challan_id):
            return r
    return None


def get_all_challans():
    records = read_all_records("Challans")
    records.sort(key=lambda r: (str(r.get("challan_date", "")), _numeric_key(r.get("challan_number")), str(r.get("challan_number", ""))))
    return records


def delete_challan(challan_id):
    with WRITE_LOCK:
        challan = get_challan(challan_id)
        if challan:
            rename_files_for_delete("CH", challan)
        delete_row_by_id("Challans", challan_id)
        for it in read_all_records("ChallanItems"):
            if str(it.get("challan_id")) == str(challan_id):
                delete_row_by_id("ChallanItems", it["id"])


def item_duplicate_in_challan(challan_id, description, qty_nos, qty_meters, exclude_item_id=None):
    for it in read_all_records("ChallanItems"):
        if (str(it.get("challan_id")) == str(challan_id) and it.get("description") == description and
            float(it.get("qty_nos") or 0) == float(qty_nos) and float(it.get("qty_meters") or 0) == float(qty_meters)):
            if exclude_item_id is not None and str(it.get("id")) == str(exclude_item_id):
                continue
            return True
    return False


def next_sr_no(challan_id):
    items = [it for it in read_all_records("ChallanItems") if str(it.get("challan_id")) == str(challan_id)]
    srs = [int(it["sr_no"]) for it in items if str(it.get("sr_no", "")).isdigit()]
    return (max(srs) + 1) if srs else 1


def insert_challan_item(challan_id, description, qty_nos, qty_meters, billable, created_by):
    with WRITE_LOCK:
        sr = next_sr_no(challan_id)
        if sr > 35:
            raise ValueError("This challan already has 35 items (maximum allowed).")
        row_id = next_id("ChallanItems")
        append_row("ChallanItems", ITEM_HEADERS, {
            "id": row_id,
            "challan_id": challan_id,
            "sr_no": sr,
            "description": description,
            "qty_nos": qty_nos,
            "qty_meters": qty_meters,
            "billable": billable,
            "created_by": created_by
        })
        return sr


def update_challan_item(item_id, description, qty_nos, qty_meters, billable):
    with WRITE_LOCK:
        item = get_challan_item(item_id)
        update_row_by_id("ChallanItems", ITEM_HEADERS, item_id, {
            "id": item_id,
            "challan_id": item["challan_id"],
            "sr_no": item["sr_no"],
            "description": description,
            "qty_nos": qty_nos,
            "qty_meters": qty_meters,
            "billable": billable,
            "created_by": item.get("created_by", "")
        })


def delete_challan_item(item_id):
    with WRITE_LOCK:
        delete_row_by_id("ChallanItems", item_id)


def get_challan_item(item_id):
    for it in read_all_records("ChallanItems"):
        if str(it.get("id")) == str(item_id):
            return it
    return None


def get_items_for_challan(challan_id):
    items = [it for it in read_all_records("ChallanItems") if str(it.get("challan_id")) == str(challan_id)]
    items.sort(key=lambda r: int(r.get("sr_no", 0)) if str(r.get("sr_no", "")).isdigit() else 0)
    return items


def get_challan_items_for_site(site_name=None):
    challans = {str(c["id"]): c for c in read_all_records("Challans")}
    items = read_all_records("ChallanItems")
    result = []
    for it in items:
        c = challans.get(str(it.get("challan_id")))
        if not c:
            continue
        if site_name and site_name != "ALL SITES" and c.get("site_name") != site_name:
            continue
        merged = dict(it)
        merged["challan_number"] = c.get("challan_number")
        merged["challan_date"] = c.get("challan_date")
        merged["site_name"] = c.get("site_name")
        merged["vehicle_number"] = c.get("vehicle_number")
        merged["driver_name"] = c.get("driver_name")
        result.append(merged)
    result.sort(key=lambda r: (
        str(r.get("challan_date", "")),
        _numeric_key(r.get("challan_number")),
        str(r.get("challan_number", "")),
        int(r.get("sr_no", 0)) if str(r.get("sr_no", "")).isdigit() else 0
    ))
    return result


def get_site_summary():
    sites = get_sites()
    purchases = read_all_records("Purchases")
    challans = read_all_records("Challans")
    summary = []
    for s in sites:
        total = sum(float(p.get("amount") or 0) for p in purchases if p.get("site_name") == s)
        n_exp = sum(1 for p in purchases if p.get("site_name") == s)
        n_ch = len({c.get("challan_number") for c in challans if c.get("site_name") == s})
        summary.append({"name": s, "total_expense": total, "n_expenses": n_exp, "n_challans": n_ch})
    return summary


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------
def build_full_export_workbook():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7A1F1F", end_color="7A1F1F", fill_type="solid")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def style_header_row(ws, row_idx=1):
        for cell in ws[row_idx]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

    def autosize(ws):
        for col_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 40)

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Purchase Entries"
    ws1.append(["Date", "Purchaser", "Vendor", "Invoice No.", "Amount", "Payment Mode", "Payment Detail",
                "Payment Date", "Site", "Challan No.", "Notes", "Entered By", "File 1", "File 2"])
    for r in get_purchases():
        ws1.append([r.get("entry_date"), r.get("purchaser"), r.get("vendor"), r.get("invoice_no"), r.get("amount"),
                    r.get("payment_mode"), r.get("payment_detail"), r.get("payment_date"), r.get("site_name"),
                    r.get("challan_number"), r.get("notes"), r.get("created_by"), r.get("file1_link"), r.get("file2_link")])
    style_header_row(ws1)
    autosize(ws1)
    ws1.freeze_panes = "A2"

    ws2 = wb.create_sheet("Challan Items")
    ws2.append(["Challan Number", "Date", "Sr No.", "Description", "Qty (Nos)", "Qty (M)", "Billable", "Site", "Vehicle", "Driver", "Entered By"])
    for r in get_challan_items_for_site():
        ws2.append([r.get("challan_number"), r.get("challan_date"), r.get("sr_no"), r.get("description"),
                    r.get("qty_nos"), r.get("qty_meters"), r.get("billable"), r.get("site_name"),
                    r.get("vehicle_number"), r.get("driver_name"), r.get("created_by")])
    style_header_row(ws2)
    autosize(ws2)
    ws2.freeze_panes = "A2"

    ws3 = wb.create_sheet("Challans")
    ws3.append(["Challan Number", "Date", "Site", "Vehicle", "Driver", "Entered By", "File 1", "File 2"])
    for c in get_all_challans():
        ws3.append([c.get("challan_number"), c.get("challan_date"), c.get("site_name"), c.get("vehicle_number"),
                    c.get("driver_name"), c.get("created_by"), c.get("file1_link"), c.get("file2_link")])
    style_header_row(ws3)
    autosize(ws3)

    ws4 = wb.create_sheet("Site Summary")
    ws4.append(["Site", "Total Expense", "No. of Entries", "No. of Challans"])
    for r in get_site_summary():
        ws4.append([r["name"], r["total_expense"], r["n_expenses"], r["n_challans"]])
    style_header_row(ws4)
    autosize(ws4)
    return wb


def build_site_export_workbook(site_name):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="7A1F1F", end_color="7A1F1F", fill_type="solid")
    sub_font = Font(bold=True, color="1F4E78", size=12)
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def style_header_row(ws, row_idx):
        for cell in ws[row_idx]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

    def autosize(ws):
        for col_cells in ws.columns:
            length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 40)

    purchases = get_purchases(site_name)
    items = get_challan_items_for_site(site_name)
    total = sum(float(p.get("amount") or 0) for p in purchases)
    n_challans = len({it.get("challan_number") for it in items})

    wb = openpyxl.Workbook()
    sh = wb.active
    sh.title = site_name[:31]
    sh.append([f"SITE: {site_name}"])
    sh["A1"].font = Font(bold=True, size=14, color="7A1F1F")
    sh.append([f"Total Purchase/Expense: {total:,.2f}    |    Unique Challans: {n_challans}"])
    sh["A2"].font = sub_font
    sh.append([])
    sh.append(["-- PURCHASE / EXPENSE ENTRIES --"])
    sh[sh.max_row][0].font = sub_font
    header_row_1 = sh.max_row + 1
    sh.append(["Date", "Purchaser", "Vendor", "Invoice No.", "Amount", "Payment Mode", "Payment Detail",
                "Payment Date", "Challan No.", "Notes", "Entered By", "File 1", "File 2"])
    for p in purchases:
        sh.append([p.get("entry_date"), p.get("purchaser"), p.get("vendor"), p.get("invoice_no"), p.get("amount"),
                   p.get("payment_mode"), p.get("payment_detail"), p.get("payment_date"), p.get("challan_number"),
                   p.get("notes"), p.get("created_by"), p.get("file1_link"), p.get("file2_link")])
    style_header_row(sh, header_row_1)
    sh.append([])
    sh.append(["-- CHALLAN ITEMS --"])
    sh[sh.max_row][0].font = sub_font
    header_row_2 = sh.max_row + 1
    sh.append(["Challan Number", "Date", "Sr No.", "Description", "Qty (Nos)", "Qty (M)", "Billable", "Vehicle", "Driver", "Entered By"])
    for it in items:
        sh.append([it.get("challan_number"), it.get("challan_date"), it.get("sr_no"), it.get("description"),
                   it.get("qty_nos"), it.get("qty_meters"), it.get("billable"), it.get("vehicle_number"),
                   it.get("driver_name"), it.get("created_by")])
    style_header_row(sh, header_row_2)
    autosize(sh)
    return wb


# ---------------------------------------------------------------------------
# HTML helpers & mobile-friendly UI
# ---------------------------------------------------------------------------
def NAV():
    role = session.get("role", "")
    # Base links visible to everyone
    nav_links = f"""
    <a href="/" style="color:#fff;text-decoration:none;margin-right:16px;">Home</a>
    <a href="/purchase" style="color:#fff;text-decoration:none;margin-right:16px;">Purchase</a>
    <a href="/challans" style="color:#fff;text-decoration:none;margin-right:16px;">Challans</a>
    <a href="/site_view" style="color:#fff;text-decoration:none;margin-right:16px;">Site-wise</a>
    """
    # Admin-only links
    if role == "Admin":
        nav_links += f"""
        <a href="/sites" style="color:#fff;text-decoration:none;margin-right:16px;">Sites</a>
        <a href="/employees" style="color:#fff;text-decoration:none;margin-right:16px;">Employees</a>
        <a href="/users" style="color:#fff;text-decoration:none;margin-right:16px;">Users</a>
        <a href="/export" style="color:#fff;text-decoration:none;margin-right:16px;">Export</a>
        """
    # User info and logout
    nav_links += f"""
    <span style="color:#ffd6d6;margin-left:auto;">
      {escape(session.get('full_name', ''))} ({escape(session.get('role', ''))}) |
      <a href="/logout" style="color:#ffd6d6;text-decoration:none;">Logout</a>
    </span>
    """
    return f"""
    <nav style="background:#1a1a1a;color:#fff;padding:0.5rem 0.75rem;display:flex;flex-wrap:wrap;gap:0.5rem;align-items:center;">
      <div style="font-weight:700;margin-right:0.5rem;">{escape(APP_TITLE)}</div>
      {nav_links}
    </nav>
    """
STYLE = """
<style>
:root {
  --bg: #f5f5f7;
  --card-bg: #ffffff;
  --text: #111111;
  --muted: #666666;
  --primary: #7A1F1F;
  --primary-contrast: #ffffff;
  --accent: #1F6F4A;
  --danger: #b00020;
  --border: #dddddd;
}

* { box-sizing: border-box; }

body {
  font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Arial, sans-serif;
  margin: 0;
  background: var(--bg);
  color: var(--text);
  line-height: 1.4;
  -webkit-text-size-adjust: 100%;
}

.container {
  padding: 0.75rem;
  max-width: 1100px;
  margin: auto;
}

/* Typography */
h2 {
  color: var(--primary);
  margin: 0.6rem 0 0.4rem 0;
  font-size: 1.35rem;
}
h3 {
  margin: 0.6rem 0 0.3rem 0;
  font-size: 1.1rem;
}
p {
  margin: 0.35rem 0;
  font-size: 0.95rem;
}

/* Cards */
.card {
  background: var(--card-bg);
  border-radius: 10px;
  box-shadow: 0 2px 6px rgba(0,0,0,0.08);
  padding: 0.85rem;
  margin-bottom: 0.85rem;
}

/* Tables */
table {
  border-collapse: collapse;
  width: 100%;
  background: var(--card-bg);
  margin-top: 0.5rem;
  font-size: 0.82rem;
}
th, td {
  border: 1px solid var(--border);
  padding: 0.45rem 0.35rem;
  text-align: left;
}
th {
  background: var(--primary);
  color: var(--primary-contrast);
  font-weight: 600;
}
tr:nth-child(even) {
  background: #fafafa;
}

/* Responsive tables */
.scroll-table {
  overflow-x: auto;
  -webkit-overflow-scrolling: touch;
}

/* Forms: smartphone-friendly vertical layout */
.card h3 {
  margin: 0 0 0.5rem 0;
  font-size: 1.15rem;
}

.card label {
  display: block;
  margin-top: 0.7rem;
  margin-bottom: 0.25rem;
  font-weight: 600;
  font-size: 0.95rem;
  color: #333;
}

.card input[type="text"],
.card input[type="number"],
.card input[type="date"],
.card select,
.card textarea {
  width: 100%;
  padding: 0.65rem 0.7rem;
  font-size: 1rem; /* prevents zoom on iOS */
  border-radius: 8px;
  border: 1px solid #ccc;
  background: #fff;
}

.card textarea {
  resize: vertical;
  min-height: 80px;
}

.card input[type="file"] {
  margin-top: 0.25rem;
  margin-bottom: 0.35rem;
  font-size: 0.9rem;
}

/* Big, touch-friendly buttons in forms */
.card button[type="submit"],
.card .btn-submit {
  width: 100%;
  padding: 0.7rem 0.8rem;
  font-size: 1.05rem;
  font-weight: 700;
  border-radius: 8px;
  min-height: 48px; /* good touch target */
  margin-top: 0.6rem;
}

/* General buttons */
button, .btn {
  display: inline-block;
  background: var(--primary);
  color: var(--primary-contrast);
  border: none;
  padding: 0.55rem 0.9rem;
  border-radius: 6px;
  cursor: pointer;
  font-size: 0.95rem;
  font-weight: 600;
  text-decoration: none;
  text-align: center;
}
button.secondary, .btn.secondary {
  background: #888;
}
button.danger, .btn.danger {
  background: var(--danger);
}
button.edit, .btn.edit {
  background: var(--accent);
}
button:disabled, .btn.disabled {
  background: #ccc;
  cursor: not-allowed;
}

/* Messages */
.msg {
  padding: 0.55rem 0.65rem;
  background: #e9fdf0;
  border: 1px solid #7ac98e;
  border-radius: 6px;
  margin-bottom: 0.55rem;
  font-size: 0.9rem;
}
.msg.error {
  background: #fde9e9;
  border-color: #e08a8a;
}

/* Grid (used in some admin forms) */
.grid {
  display: grid;
  grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
  gap: 0.5rem;
}

/* Actions */
.actions {
  white-space: nowrap;
}

/* File links */
.filelink {
  color: var(--accent);
  text-decoration: underline;
  font-weight: 600;
}

/* Navigation bar */
nav {
  background: #1a1a1a;
  color: #fff;
  padding: 0.6rem 0.75rem;
  display: flex;
  flex-wrap: wrap;
  gap: 0.5rem;
  align-items: center;
  font-size: 0.95rem;
  position: sticky;
  top: 0;
  z-index: 10;
}
nav a {
  color: #fff;
  text-decoration: none;
  padding: 0.25rem 0.35rem;
  border-radius: 4px;
}
nav a:hover {
  background: rgba(255,255,255,0.08);
}

/* Mobile tweaks */
@media (max-width: 600px) {
  .container {
    padding: 0.5rem;
  }
  h2 {
    font-size: 1.2rem;
  }
  h3 {
    font-size: 1rem;
  }
  table {
    font-size: 0.75rem;
  }
  th, td {
    padding: 0.35rem 0.25rem;
  }
  input, select, textarea {
    font-size: 1rem; /* prevents zoom on iOS */
    padding: 0.6rem 0.65rem;
  }
  button, .btn {
    font-size: 1rem;
    padding: 0.6rem 0.9rem;
    min-height: 44px; /* better touch target */
  }

  .card {
    padding: 0.7rem;
  }
  .card label {
    font-size: 0.95rem;
  }
  .card input[type="text"],
  .card input[type="number"],
  .card input[type="date"],
  .card select,
  .card textarea {
    padding: 0.7rem 0.75rem;
    font-size: 1.05rem;
  }
  .card button[type="submit"],
  .card .btn-submit {
    padding: 0.75rem 0.9rem;
    font-size: 1.1rem;
    min-height: 52px;
  }

  /* Make table actions stack a bit better */
  .actions form,
  .actions a {
    display: inline-block;
    margin-right: 0.25rem;
  }
}

/* ---------- AI Scanner Modal (fixed for desktop) ---------- */
.scanner-modal {
  display: none;
  position: fixed;
  inset: 0;
  background: rgba(0,0,0,0.85);
  z-index: 9999;
  align-items: center;
  justify-content: center;
}

.scanner-box {
  background: #fff;
  width: 92%;
  max-width: 500px;
  max-height: 85vh;
  border-radius: 12px;
  overflow: hidden;
  display: flex;
  flex-direction: column;
}

.scanner-header {
  background: #111;
  color: #fff;
  padding: 0.6rem 0.75rem;
  font-weight: 700;
  font-size: 0.95rem;
  display: flex;
  justify-content: space-between;
  align-items: center;
  flex-shrink: 0;
}

.scanner-body {
  padding: 0.6rem;
  display: flex;
  flex-direction: column;
  gap: 0.5rem;
  max-height: 70vh;
  overflow-y: auto;
}

.scanner-img-wrap {
  position: relative;
  width: 100%;
  min-height: 180px;
  max-height: 50vh;
  background: #000;
}

.scanner-img-wrap img {
  position: absolute;
  inset: 0;
  width: 100%;
  height: 100%;
  object-fit: contain;
}

.scanner-actions {
  display: flex;
  gap: 0.5rem;
  flex-shrink: 0;
}

.scanner-actions button {
  flex: 1;
  padding: 0.6rem 0.7rem;
  border-radius: 8px;
  border: none;
  font-weight: 700;
  font-size: 0.95rem;
  cursor: pointer;
  white-space: nowrap;
}

.btn-scan {
  background: #7A1F1F;
  color: #fff;
}

.btn-cancel {
  background: #888;
  color: #fff;
}

.btn-crop {
  background: #1F6F4A;
  color: #fff;
}
</style>
"""
def page(title, body, msg=None, msg_type="ok"):
    msg_html = f'<div class="msg {"error" if msg_type=="error" else ""}">{escape(msg)}</div>' if msg else ""
    return f"""<!DOCTYPE html><html><head><title>{escape(title)} - {escape(APP_TITLE)}</title>{STYLE}
    <!-- AI Document Scanner -->
    <link rel="stylesheet"
      href="https://cdnjs.cloudflare.com/ajax/libs/cropperjs/1.6.2/cropper.min.css">
    <script src="https://cdnjs.cloudflare.com/ajax/libs/cropperjs/1.6.2/cropper.min.js"></script>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/pdf-lib/1.17.1/pdf-lib.min.js"></script>
    <style>
      .scanner-modal {{
        display: none;
        position: fixed;
        inset: 0;
        background: rgba(0,0,0,0.85);
        z-index: 9999;
        align-items: center;
        justify-content: center;
      }}
      .scanner-box {{
        background: #fff;
        width: 92%;
        max-width: 500px;
        max-height: 85vh;
        border-radius: 12px;
        overflow: hidden;
        display: flex;
        flex-direction: column;
      }}
      .scanner-header {{
        background: #111;
        color: #fff;
        padding: 0.6rem 0.75rem;
        font-weight: 700;
        font-size: 0.95rem;
        display: flex;
        justify-content: space-between;
        align-items: center;
      }}
      .scanner-body {{
        padding: 0.6rem;
        display: flex;
        flex-direction: column;
        gap: 0.5rem;
      }}
      .scanner-img-wrap {{
        position: relative;
        width: 100%;
        height: 0;
        padding-bottom: 130%;
        background: #000;
      }}
      .scanner-img-wrap img {{
        position: absolute;
        inset: 0;
        width: 100%;
        height: 100%;
        object-fit: contain;
      }}
      .scanner-actions {{
        display: flex;
        gap: 0.5rem;
      }}
      .scanner-actions button {{
        flex: 1;
        padding: 0.6rem 0.7rem;
        border-radius: 8px;
        border: none;
        font-weight: 700;
        font-size: 0.95rem;
        cursor: pointer;
      }}
      .btn-scan {{
        background: #7A1F1F;
        color: #fff;
      }}
      .btn-cancel {{
        background: #888;
        color: #fff;
      }}
      .btn-crop {{
        background: #1F6F4A;
        color: #fff;
      }}
    </style>
    <script>
      let cropper = null;
      async function openScanner(targetInputId) {{
        const modal = document.getElementById('ai-scanner-modal');
        const imgEl = document.getElementById('ai-scanner-img');
        const statusEl = document.getElementById('ai-scanner-status');
        modal.style.display = 'flex';
        statusEl.textContent = 'Tap “Use camera” to capture a document.';

        const file = await new Promise((resolve) => {{
          const inner = document.createElement('input');
          inner.type = 'file';
          inner.accept = 'image/*';
          inner.capture = 'environment';
          inner.onchange = () => resolve(inner.files[0] || null);
          inner.click();
        }});

        if (!file) {{
          modal.style.display = 'none';
          return;
        }}

        const url = URL.createObjectURL(file);
        imgEl.src = url;
        imgEl.onload = () => {{
          if (cropper) cropper.destroy();
          cropper = new Cropper(imgEl, {{
            viewMode: 1,
            dragMode: 'move',
            autoCropArea: 0.9,
            responsive: true,
          }});
          statusEl.textContent = 'Adjust the crop box to match the document edges, then tap “Crop & Create PDF”.';
        }};
      }}

      async function cropAndCreatePdf(targetInputId) {{
        const modal = document.getElementById("ai-scanner-modal");
        const statusEl = document.getElementById("ai-scanner-status");
        const cropButton = modal.querySelector(".btn-crop");

        try {{
          if (!cropper) {{
            statusEl.textContent = "Please capture an image first.";
            return;
          }}

          if (!window.PDFLib || !window.PDFLib.PDFDocument) {{
            throw new Error("PDF library did not load. Please refresh the page and try again.");
          }}

          cropButton.disabled = true;
          statusEl.textContent = "Cropping and generating PDF...";

          const canvas = cropper.getCroppedCanvas({{
            maxWidth: 2000,
            maxHeight: 3000,
            imageSmoothingEnabled: true,
            imageSmoothingQuality: "high"
          }});

          if (!canvas || canvas.width === 0 || canvas.height === 0) {{
            throw new Error("The cropped image is empty.");
          }}

          const blob = await new Promise((resolve, reject) => {{
            canvas.toBlob(
              result => {{
                if (result) {{
                  resolve(result);
                }} else {{
                  reject(new Error("The browser could not create the image."));
                }}
              }},
              "image/jpeg",
              0.92
            );
          }});

          const imageBytes = new Uint8Array(await blob.arrayBuffer());
          const pdfDoc = await PDFLib.PDFDocument.create();

          const page = pdfDoc.addPage([canvas.width, canvas.height]);
          const image = await pdfDoc.embedJpg(imageBytes);

          page.drawImage(image, {{
            x: 0,
            y: 0,
            width: canvas.width,
            height: canvas.height
          }});

          const pdfBytes = await pdfDoc.save();
          const pdfBlob = new Blob([pdfBytes], {{
            type: "application/pdf"
          }});

          const pdfFile = new File(
            [pdfBlob],
            "scanned-document.pdf",
            {{ type: "application/pdf" }}
          );

          const inputEl = document.getElementById(targetInputId);

          if (!inputEl) {{
            throw new Error("The target upload field was not found.");
          }}

          const dataTransfer = new DataTransfer();
          dataTransfer.items.add(pdfFile);
          inputEl.files = dataTransfer.files;

          statusEl.textContent = "PDF ready. Submit the form to save it.";

          if (cropper) {{
            cropper.destroy();
            cropper = null;
          }}

          setTimeout(() => {{
            modal.style.display = "none";
          }}, 700);

        }} catch (error) {{
          console.error("Scanner PDF error:", error);
          statusEl.textContent =
            "Could not create PDF: " + (error.message || "Unknown error");
        }} finally {{
          cropButton.disabled = false;
        }}
      }}

      function closeScanner() {{
        const modal = document.getElementById('ai-scanner-modal');
        modal.style.display = 'none';
        if (cropper) {{
          cropper.destroy();
          cropper = null;
        }}
      }}
    </script>
    </head>
    <body>
    {NAV()}
    <div class="container">
      <h2>{escape(title)}</h2>
      {msg_html}
      {body}
    </div>

    <!-- AI Scanner Modal -->
    <div id="ai-scanner-modal" class="scanner-modal">
      <div class="scanner-box">
        <div class="scanner-header">
          <span>AI Document Scanner</span>
          <button type="button" class="btn-cancel" onclick="closeScanner()"
            style="background:transparent;border:none;color:#fff;font-weight:700;">✕</button>
        </div>
        <div class="scanner-body">
          <div class="scanner-img-wrap">
            <img id="ai-scanner-img" alt="Document to scan">
          </div>
          <div id="ai-scanner-status" style="font-size:0.9rem;color:#444;"></div>
          <div class="scanner-actions">
            <button type="button" class="btn-scan" onclick="openScanner(currentScannerTarget)">Use camera</button>
            <button type="button" class="btn-crop" onclick="cropAndCreatePdf(currentScannerTarget)">Crop & Create PDF</button>
            <button type="button" class="btn-cancel" onclick="closeScanner()">Cancel</button>
          </div>
        </div>
      </div>
    </div>

    <script>
      let currentScannerTarget = null;
    </script>
    </body></html>"""

# Global used by the in-page scanner JS
current_scanner_target = None

def file_links_html(file1, file2):
    parts = []
    if file1:
        parts.append(f'<a class="filelink" href="{escape(file1)}" target="_blank">File 1</a>')
    if file2:
        parts.append(f'<a class="filelink" href="{escape(file2)}" target="_blank">File 2</a>')
    return " | ".join(parts) if parts else "-"

# ---------------------------------------------------------------------------
# Routes: Dashboard
# ---------------------------------------------------------------------------
@app.route("/")
@login_required
def dashboard():
    summary = get_site_summary()
    rows = "".join(
        f"<tr><td>{escape(r['name'])}</td><td>{r['total_expense']:,.2f}</td>"
        f"<td>{r['n_expenses']}</td><td>{r['n_challans']}</td></tr>"
        for r in summary
    )
    body = f"""
        <div class="card">
      <p>Logged in as <b>{escape(session.get('full_name', ''))}</b> ({escape(session.get('role', ''))}).
         Employees can edit their own entries for {EDIT_WINDOW_DAYS} days; after that only the Main Admin can edit them.</p>
    </div>
    <div class="card">
      <h3>Site Summary</h3>
      <div class="scroll-table">
        <table>
          <tr><th>Site</th><th>Total Expense</th><th>Purchase Entries</th><th>Unique Challans</th></tr>
          {rows}
        </table>
      </div>
    </div>
    """
    return page("Dashboard", body)


# ---------------------------------------------------------------------------
# Routes: User Management (Main Admin only)
# ---------------------------------------------------------------------------
@app.route("/users", methods=["GET", "POST"])
@admin_required
def users_page():
    msg, msg_type = None, "ok"
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "").strip()
            full_name = request.form.get("full_name", "").strip()
            role = request.form.get("role", "Employee")
            if not username or not password or not full_name:
                msg, msg_type = "Username, password, and full name are required.", "error"
            else:
                ok, msg = add_user(username, password, full_name, role)
                msg_type = "ok" if ok else "error"
        elif action == "delete":
            ok, msg = delete_user(request.form.get("username"))

    users = get_all_users()
    rows = ""
    for u in users:
        rows += f"""<tr><td>{escape(u['username'])}</td><td>{escape(u.get('full_name', ''))}</td>
        <td>{escape(u.get('role', ''))}</td>
        <td><form method="post" onsubmit="return confirm('Delete user {escape(u['username'])}?');">
        <input type="hidden" name="action" value="delete"><input type="hidden" name="username" value="{escape(u['username'])}">
        <button type="submit" class="danger">Delete</button></form></td></tr>"""

    body = f"""<div class="card"><h3>Add New Employee Login ({len(users)}/20 used)</h3>
    <p style="font-size:0.85rem;color:#666;">Only the Main Admin role has full control. Employee accounts can add
    entries and edit only their OWN entries within {EDIT_WINDOW_DAYS} days.</p>
    <form method="post"><input type="hidden" name="action" value="add"><div class="grid">
    <div>Username<br><input type="text" name="username" required></div>
    <div>Password<br><input type="text" name="password" required></div>
    <div>Full Name<br><input type="text" name="full_name" required></div>
    <div>Role<br><select name="role"><option value="Employee">Employee</option><option value="Admin">Admin (full control)</option></select></div>
    </div><br><button type="submit">Create Login</button></form></div>
    <div class="scroll-table"><table><tr><th>Username</th><th>Full Name</th><th>Role</th><th></th></tr>{rows}</table></div>"""
    return page("Manage Employee Logins", body, msg, msg_type)


# ---------------------------------------------------------------------------
# Routes: Sites / Employees CRUD (Admin only)
# ---------------------------------------------------------------------------
@app.route("/sites", methods=["GET", "POST"])
@admin_required
def sites_page():
    msg, msg_type = None, "ok"
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            ok, msg = add_site(request.form.get("name", "").strip())
            msg_type = "ok" if ok else "error"
        elif action == "rename":
            rename_site(request.form.get("old_name"), request.form.get("new_name", "").strip())
            msg = "Site renamed everywhere it is used."
        elif action == "delete":
            ok, msg = delete_site(request.form.get("name"), force=request.form.get("force") == "1")
            msg_type = "ok" if ok else "error"

    sites = get_sites()
    rows = ""
    for s in sites:
        rows += f"""<tr><td>{escape(s)}</td>
          <td><form class="inline" method="post"><input type="hidden" name="action" value="rename">
          <input type="hidden" name="old_name" value="{escape(s)}"><input type="text" name="new_name" placeholder="New name" required>
          <button type="submit" class="edit">Rename</button></form></td>
          <td class="actions"><form class="inline" method="post" onsubmit="return confirm('Delete {escape(s)}?');">
          <input type="hidden" name="action" value="delete"><input type="hidden" name="name" value="{escape(s)}">
          <button type="submit" class="danger">Delete</button></form>
          <form class="inline" method="post" onsubmit="return confirm('Force delete {escape(s)}?');">
          <input type="hidden" name="action" value="delete"><input type="hidden" name="name" value="{escape(s)}">
          <input type="hidden" name="force" value="1"><button type="submit" class="danger">Force Delete</button></form></td></tr>"""
    body = f"""<div class="card"><h3>Add New Site ({len(sites)}/{MAX_SITES} used)</h3>
    <form method="post"><input type="hidden" name="action" value="add">
    <input type="text" name="name" placeholder="New site name" required>
    <button type="submit">Add Site</button></form></div>
    <div class="scroll-table"><table><tr><th>Site Name</th><th>Rename</th><th>Delete</th></tr>{rows}</table></div>"""
    return page("Manage Sites", body, msg, msg_type)


@app.route("/employees", methods=["GET", "POST"])
@admin_required
def employees_page():
    msg, msg_type = None, "ok"
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            ok, msg = add_employee(request.form.get("name", "").strip())
            msg_type = "ok" if ok else "error"
        elif action == "rename":
            rename_employee(request.form.get("old_name"), request.form.get("new_name", "").strip())
            msg = "Employee renamed."
        elif action == "delete":
            ok, msg = delete_employee(request.form.get("name"))

    employees = get_employees()
    rows = ""
    for e in employees:
        rows += f"""<tr><td>{escape(e)}</td>
          <td><form class="inline" method="post"><input type="hidden" name="action" value="rename">
          <input type="hidden" name="old_name" value="{escape(e)}"><input type="text" name="new_name" placeholder="New name" required>
          <button type="submit" class="edit">Rename</button></form></td>
          <td><form class="inline" method="post" onsubmit="return confirm('Delete {escape(e)}?');">
          <input type="hidden" name="action" value="delete"><input type="hidden" name="name" value="{escape(e)}">
          <button type="submit" class="danger">Delete</button></form></td></tr>"""
    body = f"""<div class="card"><h3>Add New Employee ({len(employees)}/{MAX_EMPLOYEES} used)</h3>
    <form method="post"><input type="hidden" name="action" value="add">
    <input type="text" name="name" placeholder="New employee name" required>
    <button type="submit">Add Employee</button></form></div>
    <div class="scroll-table"><table><tr><th>Employee Name</th><th>Rename</th><th>Delete</th></tr>{rows}</table></div>"""
    return page("Manage Employees", body, msg, msg_type)

  # ---------------------------------------------------------------------------
# Routes: Purchase (with file upload + permission-gated editing)
# ---------------------------------------------------------------------------
def render_purchase_form_and_table():
    sites_opts = "".join(f'<option value="{escape(s)}">{escape(s)}</option>' for s in get_sites())
    emp_opts = "".join(f'<option value="{escape(e)}">{escape(e)}</option>' for e in get_employees())
    pm_opts = "".join(f'<option value="{p}">{p}</option>' for p in PAYMENT_MODES)
    today = datetime.today().strftime("%Y-%m-%d")

    # Smartphone-friendly vertical form
    form = f"""<div class="card"><h3>New Entry</h3>
    <form method="post" enctype="multipart/form-data">
    <input type="hidden" name="action" value="add">

    <label>Date</label>
    <input type="date" name="entry_date" value="{today}" required>

    <label>Purchaser</label>
    <select name="purchaser">{emp_opts}</select>

    <label>Vendor/Details</label>
    <input type="text" name="vendor" placeholder="Vendor name / details">

    <label>Invoice No.</label>
    <input type="text" name="invoice_no" placeholder="Invoice number">

    <label>Amount (with GST)</label>
    <input type="number" step="0.01" name="amount" placeholder="0.00">

    <label>Payment Mode</label>
    <select name="payment_mode">{pm_opts}</select>

    <label>Payment Detail</label>
    <input type="text" name="payment_detail" placeholder="e.g. UTX ID, Cheque No.">

    <label>Payment Date</label>
    <input type="date" name="payment_date" value="{today}">

    <label>Site Name</label>
    <select name="site_name">{sites_opts}</select>

    <label>Challan Number</label>
    <input type="text" name="challan_number" placeholder="Challan number">

    <label>Notes / Remarks</label>
    <textarea name="notes" rows="3" placeholder="Any additional notes"></textarea>

    <label>Attach file 1 (JPG/JPEG/PDF)</label>
    <input type="file" id="purchase_file1" name="attachments" accept=".jpg,.jpeg,.pdf">
    <button type="button"
            onclick="currentScannerTarget='purchase_file1'; openScanner('purchase_file1');"
            class="btn"
            style="width:100%;margin-top:0.35rem;">
      Scan document with AI
    </button>

    <label>Attach file 2 (JPG/JPEG/PDF)</label>
    <input type="file" id="purchase_file2" name="attachments" accept=".jpg,.jpeg,.pdf">
    <button type="button"
            onclick="currentScannerTarget='purchase_file2'; openScanner('purchase_file2');"
            class="btn"
            style="width:100%;margin-top:0.35rem;">
      Scan document with AI
    </button>

    <div style="margin-top:0.6rem;">
      <button type="submit" style="width:100%;font-size:1.05rem;">Add Entry</button>
    </div>
    </form></div>"""
    
    rows = ""
    for r in get_purchases():
        allowed, _ = can_modify_entry(r.get("created_by", ""), r.get("created_at", ""))
        edit_btn = f'<a href="/purchase/{r.get("id")}/edit"><button type="button" class="edit">Edit</button></a>' if allowed else '<button type="button" disabled>Locked</button>'
        del_btn = (f'<form class="inline" method="post" onsubmit="return confirm(\'Delete this entry?\');">'
                   f'<input type="hidden" name="action" value="delete"><input type="hidden" name="id" value="{r.get("id")}">'
                   f'<button type="submit" class="danger">Delete</button></form>') if allowed else ""
        rows += f"""<tr><td>{r.get('id')}</td><td>{escape(str(r.get('entry_date') or ''))}</td>
        <td>{escape(str(r.get('purchaser') or ''))}</td><td>{escape(str(r.get('vendor') or ''))}</td>
        <td>{escape(str(r.get('invoice_no') or ''))}</td><td>{float(r.get('amount') or 0):,.2f}</td>
        <td>{escape(str(r.get('payment_mode') or ''))}</td><td>{escape(str(r.get('payment_detail') or ''))}</td>
        <td>{escape(str(r.get('payment_date') or ''))}</td><td>{escape(str(r.get('site_name') or ''))}</td>
        <td>{escape(str(r.get('challan_number') or ''))}</td><td>{escape(str(r.get('notes') or ''))}</td>
        <td><b>{escape(str(r.get('created_by') or ''))}</b></td>
        <td>{file_links_html(r.get('file1_link'), r.get('file2_link'))}</td>
        <td class="actions">{edit_btn} {del_btn}</td></tr>"""
    table = f"""<div class="scroll-table"><table><tr><th>ID</th><th>Date</th><th>Purchaser</th><th>Vendor</th><th>Invoice</th>
    <th>Amount</th><th>Mode</th><th>Detail</th><th>Pay Date</th><th>Site</th><th>Challan</th><th>Notes</th>
    <th>Entered By</th><th>Files</th><th></th></tr>{rows}</table></div>"""
    return form + table


@app.route("/purchase", methods=["GET", "POST"])
@login_required
def purchase_page():
    msg, msg_type = None, "ok"
    if request.method == "POST":
        action = request.form.get("action")
        if action == "delete":
            row_id = request.form.get("id")
            row = get_purchase_row(row_id)
            allowed, reason = can_modify_entry(row.get("created_by", ""), row.get("created_at", "")) if row else (False, "Not found.")
            if not allowed:
                msg, msg_type = reason, "error"
            else:
                delete_purchase(row_id)
                msg = "Entry deleted."
        else:
            f = request.form
            date_ = f.get("entry_date", "").strip()
            site = f.get("site_name", "").strip()
            try:
                amount = float(f.get("amount") or 0)
            except ValueError:
                amount = 0.0
            vendor = f.get("vendor", "").strip()
            invoice_no = f.get("invoice_no", "").strip()
            vendor = vendor.strip()
            if not date_ or not site:
                msg, msg_type = "Date and Site are required.", "error"
            elif not vendor:
                msg, msg_type = "Vendor name is required.", "error"
            elif amount <= 0:
                msg, msg_type = "Amount must be greater than 0.", "error"
            else:
                files = request.files.getlist("attachments")
                ok, ferr = validate_uploaded_files([x for x in files if x and x.filename])
                if not ok:
                    msg, msg_type = ferr, "error"
                    return page("Purchase Entry", render_purchase_form_and_table(), msg, msg_type)
                if invoice_no and check_duplicate_purchase(date_, vendor, invoice_no, amount) and f.get("confirm_dup") != "1":
                    msg, msg_type = "Possible duplicate found. Submit again to confirm and add anyway.", "error"
                    return page("Purchase Entry", render_purchase_form_and_table(), msg, msg_type)
                new_id = insert_purchase({
                    "entry_date": date_, "purchaser": f.get("purchaser", "").strip(), "vendor": vendor,
                    "invoice_no": invoice_no, "amount": amount, "payment_mode": f.get("payment_mode", "").strip(),
                    "payment_detail": f.get("payment_detail", "").strip(), "payment_date": f.get("payment_date", "").strip(),
                    "site_name": site, "challan_number": f.get("challan_number", "").strip(),
                    "notes": f.get("notes", "").strip()}, current_user(), files=files)
                msg = "Entry added."
    return page("Purchase Entry", render_purchase_form_and_table(), msg, msg_type)


@app.route("/purchase/<row_id>/edit", methods=["GET", "POST"])
@login_required
def purchase_edit(row_id):
    row = get_purchase_row(row_id)
    if not row:
        return redirect(url_for("purchase_page"))
    allowed, reason = can_modify_entry(row.get("created_by", ""), row.get("created_at", ""))
    if not allowed:
        return page("Cannot Edit", f"<p>{escape(reason)}</p><p><a href='/purchase'>Back</a></p>", reason, "error")

    msg, msg_type = None, "ok"
    if request.method == "POST":
        f = request.form
        date_ = f.get("entry_date", "").strip()
        site = f.get("site_name", "").strip()
        try:
            amount = float(f.get("amount") or 0)
        except ValueError:
            amount = 0.0
        if not date_ or not site:
            msg, msg_type = "Date and Site are required.", "error"
        else:
            files = request.files.getlist("attachments")
            real_files = [x for x in files if x and x.filename]
            ok, ferr = validate_uploaded_files(real_files)
            if not ok:
                msg, msg_type = ferr, "error"
            else:
                record = {
                    "entry_date": date_, "purchaser": f.get("purchaser", "").strip(), "vendor": f.get("vendor", "").strip(),
                    "invoice_no": f.get("invoice_no", "").strip(), "amount": amount, "payment_mode": f.get("payment_mode", "").strip(),
                    "payment_detail": f.get("payment_detail", "").strip(), "payment_date": f.get("payment_date", "").strip(),
                    "site_name": site, "challan_number": f.get("challan_number", "").strip(), "notes": f.get("notes", "").strip()
                }
                if real_files:
                    f1, f2, uerr = process_file_uploads_with_site(real_files, "PUR", str(row_id), record["site_name"])
                    if uerr:
                        msg, msg_type = f"Could not update files: {uerr}", "error"
                        return page(f"Edit Purchase Entry #{row_id}", "", msg, msg_type)
                    if f1:
                        record["file1_link"] = f1
                    if f2:
                        record["file2_link"] = f2
                update_purchase(row_id, record, files=files if real_files else None)
                return redirect(url_for("purchase_page"))
        row = get_purchase_row(row_id)

    sites_opts = "".join(f'<option value="{escape(s)}" {"selected" if s == row.get("site_name") else ""}>{escape(s)}</option>' for s in get_sites())
    emp_opts = "".join(f'<option value="{escape(e)}" {"selected" if e == row.get("purchaser") else ""}>{escape(e)}</option>' for e in get_employees())
    pm_opts = "".join(f'<option value="{p}" {"selected" if p == row.get("payment_mode") else ""}>{p}</option>' for p in PAYMENT_MODES)

    body = f"""<div class="card">
    <p><b>Originally entered by:</b> {escape(str(row.get('created_by') or 'unknown'))} |
    <b>Current files:</b> {file_links_html(row.get('file1_link'), row.get('file2_link'))}</p>

    <form method="post" enctype="multipart/form-data">

    <label>Date</label>
    <input type="date" name="entry_date" value="{escape(str(row.get('entry_date') or ''))}" required>

    <label>Purchaser</label>
    <select name="purchaser">{emp_opts}</select>

    <label>Vendor/Details</label>
    <input type="text" name="vendor" value="{escape(str(row.get('vendor') or ''))}">

    <label>Invoice No.</label>
    <input type="text" name="invoice_no" value="{escape(str(row.get('invoice_no') or ''))}">

    <label>Amount (with GST)</label>
    <input type="number" step="0.01" name="amount" value="{row.get('amount')}">

    <label>Payment Mode</label>
    <select name="payment_mode">{pm_opts}</select>

    <label>Payment Detail</label>
    <input type="text" name="payment_detail" value="{escape(str(row.get('payment_detail') or ''))}">

    <label>Payment Date</label>
    <input type="date" name="payment_date" value="{escape(str(row.get('payment_date') or ''))}">

    <label>Site Name</label>
    <select name="site_name">{sites_opts}</select>

    <label>Challan Number</label>
    <input type="text" name="challan_number" value="{escape(str(row.get('challan_number') or ''))}">

    <label>Notes / Remarks</label>
    <textarea name="notes" rows="3">{escape(str(row.get('notes') or ''))}</textarea>

    <label>Replace file 1 (JPG/JPEG/PDF)</label>
    <input type="file" id="purchase_edit_file1" name="attachments" accept=".jpg,.jpeg,.pdf">
    <button type="button"
            onclick="currentScannerTarget='purchase_edit_file1'; openScanner('purchase_edit_file1');"
            class="btn"
            style="width:100%;margin-top:0.35rem;">
      Scan document with AI
    </button>

    <label>Replace file 2 (JPG/JPEG/PDF)</label>
    <input type="file" id="purchase_edit_file2" name="attachments" accept=".jpg,.jpeg,.pdf">
    <button type="button"
            onclick="currentScannerTarget='purchase_edit_file2'; openScanner('purchase_edit_file2');"
            class="btn"
            style="width:100%;margin-top:0.35rem;">
      Scan document with AI
    </button>

    <div style="margin-top:0.6rem; display:flex; gap:0.5rem;">
      <button type="submit" style="flex:1;font-size:1.05rem;">Save Changes</button>
      <a href="/purchase" style="flex:1;text-decoration:none;">
        <button type="button" class="secondary" style="width:100%;font-size:1.05rem;">Cancel</button>
      </a>
    </div>
    </form></div>"""

  # ---------------------------------------------------------------------------
# Routes: Challans (with file upload + permission-gated editing)
# ---------------------------------------------------------------------------
@app.route("/challans", methods=["GET", "POST"])
@login_required
def challans_page():
    msg, msg_type = None, "ok"

    # Define sites_opts here so it exists for both GET and POST
    sites_opts = "".join(
        f'<option value="{escape(s)}">{escape(s)}</option>' for s in get_sites()
    )

    if request.method == "POST":
        f = request.form
        challan_no = f.get("challan_number", "").strip()
        challan_date = f.get("challan_date", "").strip()
        site = f.get("site_name", "").strip()
        ...
        if not challan_no or not challan_date or not site:
            msg, msg_type = "Challan Number, Date, and Site are required.", "error"
        else:
            mismatch = challan_date_mismatch(challan_no, challan_date)
            if mismatch:
                msg, msg_type = f"Challan '{challan_no}' already exists with date {mismatch}.", "error"
            else:
                files = request.files.getlist("attachments")
                real_files = [x for x in files if x and x.filename]
                ok, ferr = validate_uploaded_files(real_files)
                if not ok:
                    msg, msg_type = ferr, "error"
                else:
                    cid = get_or_create_challan(
                        challan_no, challan_date, site,
                        f.get("vehicle_number", "").strip(),
                        f.get("driver_name", "").strip(),
                        current_user()
                    )
                    if real_files:
                        f1, f2, uerr = process_file_uploads_with_site(real_files, "CH", challan_no, site)
                        if uerr:
                            msg, msg_type = uerr, "error"
                        else:
                            update_challan_header(
                                cid, challan_date, site,
                                f.get("vehicle_number", "").strip(),
                                f.get("driver_name", "").strip(),
                                f1, f2, files=real_files
                            )
                    return redirect(url_for("challan_detail", challan_id=cid))

    sites_opts = "".join(f'<option value="{escape(s)}">{escape(s)}</option>' for s in get_sites())
    today = datetime.today().strftime("%Y-%m-%d")

    form = f"""<div class="card"><h3>Start / Load a Challan</h3>
    <form method="post" enctype="multipart/form-data">

    <label>Challan Number</label>
    <input type="text" name="challan_number" placeholder="Challan number" required>

    <label>Date</label>
    <input type="date" name="challan_date" value="{today}" required>

    <label>Site</label>
    <select name="site_name">{sites_opts}</select>

    <label>Vehicle Number</label>
    <input type="text" name="vehicle_number" placeholder="Vehicle number">

    <label>Driver Name</label>
    <input type="text" name="driver_name" placeholder="Driver name">

    <label>Attach file 1 (JPG/JPEG/PDF)</label>
    <input type="file" id="challan_file1" name="attachments" accept=".jpg,.jpeg,.pdf">
    <button type="button"
            onclick="currentScannerTarget='challan_file1'; openScanner('challan_file1');"
            class="btn"
            style="width:100%;margin-top:0.35rem;">
      Scan document with AI
    </button>

    <label>Attach file 2 (JPG/JPEG/PDF)</label>
    <input type="file" id="challan_file2" name="attachments" accept=".jpg,.jpeg,.pdf">
    <button type="button"
            onclick="currentScannerTarget='challan_file2'; openScanner('challan_file2');"
            class="btn"
            style="width:100%;margin-top:0.35rem;">
      Scan document with AI
    </button>

    <div style="margin-top:0.6rem;">
      <button type="submit" style="width:100%;font-size:1.05rem;">Start / Load Challan</button>
    </div>
    </form></div>"""

    rows = ""
    for c in get_all_challans():
        n_items = len(get_items_for_challan(c["id"]))
        allowed, _ = can_modify_entry(c.get("created_by", ""), c.get("created_at", ""))
        edit_btn = f'<a href="/challans/{c["id"]}/edit"><button type="button" class="edit">Edit</button></a>' if allowed else '<button type="button" disabled>Locked</button>'
        del_btn = (f'<form class="inline" method="post" action="/challans/{c["id"]}/delete" onsubmit="return confirm(\'Delete this challan and items?\');">'
                   f'<button type="submit" class="danger">Delete</button></form>') if allowed else ""
        rows += f"""<tr><td><a href="/challans/{c['id']}">{escape(str(c['challan_number']))}</a></td>
        <td>{escape(str(c['challan_date']))}</td><td>{escape(str(c['site_name']))}</td>
        <td>{escape(str(c.get('vehicle_number') or ''))}</td><td>{escape(str(c.get('driver_name') or ''))}</td><td>{n_items}</td>
        <td><b>{escape(str(c.get('created_by') or ''))}</b></td>
        <td>{file_links_html(c.get('file1_link'), c.get('file2_link'))}</td>
        <td class="actions">{edit_btn} {del_btn}</td></tr>"""
    table = f"""<div class="scroll-table"><table><tr><th>Challan No.</th><th>Date</th><th>Site</th><th>Vehicle</th><th>Driver</th>
    <th>Items</th><th>Entered By</th><th>Files</th><th></th></tr>{rows}</table></div>"""
    return page("Challans", form + table, msg, msg_type)


@app.route("/challans/<challan_id>/delete", methods=["POST"])
@login_required
def challan_delete_route(challan_id):
    challan = get_challan(challan_id)
    if challan:
        allowed, reason = can_modify_entry(challan.get("created_by", ""), challan.get("created_at", ""))
        if allowed:
            delete_challan(challan_id)
    return redirect(url_for("challans_page"))


@app.route("/challans/<challan_id>/edit", methods=["GET", "POST"])
@login_required
def challan_edit(challan_id):
    challan = get_challan(challan_id)
    if not challan:
        return redirect(url_for("challans_page"))
    allowed, reason = can_modify_entry(challan.get("created_by", ""), challan.get("created_at", ""))
    if not allowed:
        return page("Cannot Edit", f"<p>{escape(reason)}</p><p><a href='/challans'>Back</a></p>", reason, "error")

    msg, msg_type = None, "ok"
    if request.method == "POST":
        f = request.form
        challan_date = f.get("challan_date", "").strip()
        site = f.get("site_name", "").strip()
        mismatch = challan_date_mismatch(challan["challan_number"], challan_date, exclude_id=challan_id)
        if mismatch:
            msg, msg_type = f"Another challan with the same number already uses date {mismatch}.", "error"
        else:
            files = request.files.getlist("attachments")
            real_files = [x for x in files if x and x.filename]
            ok, ferr = validate_uploaded_files(real_files)
            if not ok:
                msg, msg_type = ferr, "error"
            else:
                f1, f2 = None, None
                if real_files:
                    f1, f2, uerr = process_file_uploads_with_site(real_files, "CH", challan["challan_number"], site)
                    if uerr:
                        msg, msg_type = uerr, "error"
                        challan = get_challan(challan_id)
                        return page(f"Edit Challan {challan['challan_number']}", "", msg, msg_type)
                update_challan_header(
                    challan_id, challan_date, site,
                    f.get("vehicle_number", "").strip(),
                    f.get("driver_name", "").strip(),
                    f1, f2, files=real_files
                )
                return redirect(url_for("challan_detail", challan_id=challan_id))
        challan = get_challan(challan_id)

    sites_opts = "".join(f'<option value="{escape(s)}" {"selected" if s == challan["site_name"] else ""}>{escape(s)}</option>' for s in get_sites())

    body = f"""<div class="card">
    <p><b>Challan Number:</b> {escape(str(challan['challan_number']))} (fixed) |
    <b>Originally entered by:</b> {escape(str(challan.get('created_by') or 'unknown'))} |
    <b>Current files:</b> {file_links_html(challan.get('file1_link'), challan.get('file2_link'))}</p>

    <form method="post" enctype="multipart/form-data">

    <label>Date</label>
    <input type="date" name="challan_date" value="{escape(str(challan['challan_date']))}" required>

    <label>Site</label>
    <select name="site_name">{sites_opts}</select>

    <label>Vehicle Number</label>
    <input type="text" name="vehicle_number" value="{escape(str(challan.get('vehicle_number') or ''))}">

    <label>Driver Name</label>
    <input type="text" name="driver_name" value="{escape(str(challan.get('driver_name') or ''))}">

    <label>Replace file 1 (JPG/JPEG/PDF)</label>
    <input type="file" id="challan_edit_file1" name="attachments" accept=".jpg,.jpeg,.pdf">
    <button type="button"
            onclick="currentScannerTarget='challan_edit_file1'; openScanner('challan_edit_file1');"
            class="btn"
            style="width:100%;margin-top:0.35rem;">
      Scan document with AI
    </button>

    <label>Replace file 2 (JPG/JPEG/PDF)</label>
    <input type="file" id="challan_edit_file2" name="attachments" accept=".jpg,.jpeg,.pdf">
    <button type="button"
            onclick="currentScannerTarget='challan_edit_file2'; openScanner('challan_edit_file2');"
            class="btn"
            style="width:100%;margin-top:0.35rem;">
      Scan document with AI
    </button>

    <div style="margin-top:0.6rem; display:flex; gap:0.5rem;">
      <button type="submit" style="flex:1;font-size:1.05rem;">Save Changes</button>
      <a href="/challans/{challan_id}" style="flex:1;text-decoration:none;">
        <button type="button" class="secondary" style="width:100%;font-size:1.05rem;">Cancel</button>
      </a>
    </div>
    </form></div>"""
    return page(f"Edit Challan {challan['challan_number']}", body, msg, msg_type)


# ---------------------------------------------------------------------------
# Routes: Challan detail (items) & item edit
# ---------------------------------------------------------------------------
@app.route("/challans/<challan_id>", methods=["GET", "POST"])
@login_required
def challan_detail(challan_id):
    challan = get_challan(challan_id)
    if not challan:
        return redirect(url_for("challans_page"))
    challan_allowed, challan_reason = can_modify_entry(challan.get("created_by", ""), challan.get("created_at", ""))
    msg, msg_type = None, "ok"
    if request.method == "POST":
        action = request.form.get("action")
        if action == "delete_item":
            item_id = request.form.get("item_id")
            item = get_challan_item(item_id)
            item_allowed, item_reason = can_modify_entry(item.get("created_by", ""), challan.get("created_at", "")) if item else (False, "Not found")
            if not item_allowed:
                msg, msg_type = item_reason, "error"
            else:
                delete_challan_item(item_id)
                msg = "Item deleted."
        else:
            if not challan_allowed:
                msg, msg_type = challan_reason, "error"
            else:
                f = request.form
                desc = f.get("description", "").strip()
                try:
                    qty_nos = float(f.get("qty_nos") or 0)
                    qty_m = float(f.get("qty_meters") or 0)
                except ValueError:
                    qty_nos, qty_m = 0.0, 0.0
                billable = f.get("billable", "").strip()
                if not desc or billable not in BILLABLE_OPTIONS:
                    msg, msg_type = "Description and Billable are required.", "error"
                elif item_duplicate_in_challan(challan_id, desc, qty_nos, qty_m):
                    msg, msg_type = "An item with the same Description and Quantity already exists in THIS challan.", "error"
                else:
                    try:
                        insert_challan_item(challan_id, desc, qty_nos, qty_m, billable, current_user())
                        msg = "Item added."
                    except ValueError as e:
                        msg, msg_type = str(e), "error"

    items = get_items_for_challan(challan_id)
    rows = ""
    for it in items:
        item_allowed, _ = can_modify_entry(it.get("created_by", ""), challan.get("created_at", ""))
        edit_btn = f'<a href="/challans/{challan_id}/items/{it["id"]}/edit"><button type="button" class="edit">Edit</button></a>' if item_allowed else '<button type="button" disabled>Locked</button>'
        del_btn = (f'<form class="inline" method="post" onsubmit="return confirm(\'Delete this item?\');">'
                   f'<input type="hidden" name="action" value="delete_item"><input type="hidden" name="item_id" value="{it["id"]}">'
                   f'<button type="submit" class="danger">Delete</button></form>') if item_allowed else ""
        rows += f"""<tr><td>{it['sr_no']}</td><td>{escape(str(it['description']))}</td><td>{it['qty_nos']}</td>
        <td>{it['qty_meters']}</td><td>{escape(str(it['billable']))}</td><td><b>{escape(str(it.get('created_by') or ''))}</b></td>
        <td class="actions">{edit_btn} {del_btn}</td></tr>"""

    bill_opts = "".join(f'<option value="{b}">{b}</option>' for b in BILLABLE_OPTIONS)
    add_item_form = f"""<div class="card"><h3>Add Item</h3><form method="post"><div class="grid">
    <div>Description<br><input type="text" name="description" required></div>
    <div>Qty (Nos)<br><input type="number" step="0.01" name="qty_nos"></div>
    <div>Qty (Meters)<br><input type="number" step="0.01" name="qty_meters"></div>
    <div>Billable<br><select name="billable">{bill_opts}</select></div></div>
    <br><button type="submit">Add Item</button></form></div>""" if challan_allowed else \
    f"""<div class="card"><p style="color:#b00020;">{escape(challan_reason)} (No new items can be added to this challan.)</p></div>"""

    body = f"""<div class="card"><p><b>Challan:</b> {escape(str(challan['challan_number']))} &nbsp;
    <b>Date:</b> {escape(str(challan['challan_date']))} &nbsp; <b>Site:</b> {escape(str(challan['site_name']))} &nbsp;
    <b>Vehicle:</b> {escape(str(challan.get('vehicle_number') or '-'))} &nbsp;
    <b>Driver:</b> {escape(str(challan.get('driver_name') or '-'))} &nbsp; ({len(items)}/35 items) &nbsp;
    <b>Files:</b> {file_links_html(challan.get('file1_link'), challan.get('file2_link'))} &nbsp;
    {"<a href='/challans/"+str(challan_id)+"/edit'><button type='button' class='edit'>Edit Challan Header</button></a>" if challan_allowed else ""}</p></div>
    {add_item_form}
    <div class="scroll-table"><table><tr><th>Sr No.</th><th>Description</th><th>Qty(Nos)</th><th>Qty(M)</th><th>Billable</th><th>Entered By</th><th></th></tr>{rows}</table></div>
    <p><a href="/challans">&larr; Back to all challans</a></p>"""
    return page(f"Challan {challan['challan_number']}", body, msg, msg_type)


@app.route("/challans/<challan_id>/items/<item_id>/edit", methods=["GET", "POST"])
@login_required
def challan_item_edit(challan_id, item_id):
    item = get_challan_item(item_id)
    challan = get_challan(challan_id)
    if not item or not challan:
        return redirect(url_for("challan_detail", challan_id=challan_id))
    allowed, reason = can_modify_entry(item.get("created_by", ""), challan.get("created_at", ""))
    if not allowed:
        return page("Cannot Edit", f"<p>{escape(reason)}</p><p><a href='/challans/{challan_id}'>Back</a></p>", reason, "error")

    msg, msg_type = None, "ok"
    if request.method == "POST":
        f = request.form
        desc = f.get("description", "").strip()
        try:
            qty_nos = float(f.get("qty_nos") or 0)
            qty_m = float(f.get("qty_meters") or 0)
        except ValueError:
            qty_nos, qty_m = 0.0, 0.0
        billable = f.get("billable", "").strip()
        if not desc or billable not in BILLABLE_OPTIONS:
            msg, msg_type = "Description and Billable are required.", "error"
        elif item_duplicate_in_challan(challan_id, desc, qty_nos, qty_m, exclude_item_id=item_id):
            msg, msg_type = "Another item with the same Description and Quantity already exists in this challan.", "error"
        else:
            update_challan_item(item_id, desc, qty_nos, qty_m, billable)
            return redirect(url_for("challan_detail", challan_id=challan_id))
        item = get_challan_item(item_id)

    bill_opts = "".join(f'<option value="{b}" {"selected" if b == item["billable"] else ""}>{b}</option>' for b in BILLABLE_OPTIONS)
    body = f"""<div class="card"><p><b>Sr No.:</b> {item['sr_no']} (fixed) |
    <b>Originally entered by:</b> {escape(str(item.get('created_by') or 'unknown'))}</p><form method="post"><div class="grid">
    <div>Description<br><input type="text" name="description" value="{escape(str(item['description']))}" required></div>
    <div>Qty (Nos)<br><input type="number" step="0.01" name="qty_nos" value="{item['qty_nos']}"></div>
    <div>Qty (Meters)<br><input type="number" step="0.01" name="qty_meters" value="{item['qty_meters']}"></div>
    <div>Billable<br><select name="billable">{bill_opts}</select></div></div>
    <br><button type="submit">Save Changes</button>
    <a href="/challans/{challan_id}"><button type="button" class="secondary">Cancel</button></a></form></div>"""
    return page(f"Edit Item (Sr No. {item['sr_no']})", body, msg, msg_type)


# ---------------------------------------------------------------------------
# Routes: Site-wise View & Export
# ---------------------------------------------------------------------------
@app.route("/site_view")
@login_required
def site_view():
    site = request.args.get("site", "ALL SITES")
    sites_opts = "".join(f'<option value="{escape(s)}" {"selected" if s == site else ""}>{escape(s)}</option>' for s in ["ALL SITES"] + get_sites())
    pc_rows = "".join(
        f"<tr><td>{escape(str(r.get('entry_date') or ''))}</td><td>{escape(str(r.get('purchaser') or ''))}</td>"
        f"<td>{escape(str(r.get('vendor') or ''))}</td><td>{escape(str(r.get('invoice_no') or ''))}</td>"
        f"<td>{float(r.get('amount') or 0):,.2f}</td><td>{escape(str(r.get('site_name') or ''))}</td>"
        f"<td>{escape(str(r.get('challan_number') or ''))}</td><td>{escape(str(r.get('notes') or ''))}</td>"
        f"<td>{escape(str(r.get('created_by') or ''))}</td><td>{file_links_html(r.get('file1_link'), r.get('file2_link'))}</td></tr>"
        for r in get_purchases(site)
    )
    ci_rows = "".join(
        f"<tr><td>{escape(str(r.get('challan_number')))}</td><td>{escape(str(r.get('challan_date')))}</td>"
        f"<td>{r.get('sr_no')}</td><td>{escape(str(r.get('description')))}</td><td>{r.get('qty_nos')}</td>"
        f"<td>{r.get('qty_meters')}</td><td>{escape(str(r.get('billable')))}</td><td>{escape(str(r.get('site_name')))}</td>"
        f"<td>{escape(str(r.get('created_by') or ''))}</td></tr>"
        for r in get_challan_items_for_site(site)
    )
    total = sum(float(r.get("amount") or 0) for r in get_purchases(site)) if site != "ALL SITES" else None
    total_html = f"<p><b>Total Expense for {escape(site)}: {total:,.2f}</b></p>" if total is not None else ""

    # Export button only for Admin
    export_btn = ""
    if is_admin() and site != "ALL SITES":
        export_btn = f'<a href="/export/site/{escape(site)}"><button type="button">Download {escape(site)} Report (Excel)</button></a>'

    body = f"""<form method="get"><label>Select Site: </label>
    <select name="site" onchange="this.form.submit()">{sites_opts}</select></form>{total_html}{export_btn}
    <h3>Purchase Entries</h3>
    <div class="scroll-table"><table><tr><th>Date</th><th>Purchaser</th><th>Vendor</th><th>Invoice</th>
    <th>Amount</th><th>Site</th><th>Challan</th><th>Notes</th><th>Entered By</th><th>Files</th></tr>{pc_rows}</table></div>
    <h3>Challan Items</h3>
    <div class="scroll-table"><table><tr><th>Challan No.</th><th>Date</th><th>Sr No.</th><th>Description</th>
    <th>Qty(Nos)</th><th>Qty(M)</th><th>Billable</th><th>Site</th><th>Entered By</th></tr>{ci_rows}</table></div>"""
    return page("Site-wise View", body)


@app.route("/export")
@admin_required
def export_route():
    wb = build_full_export_workbook()
    tmp_path = "/tmp/Fire_Safety_Full_Export.xlsx"
    wb.save(tmp_path)
    return send_file(tmp_path, as_attachment=True, download_name="Fire_Safety_Full_Export.xlsx")


@app.route("/export/site/<site_name>")
@admin_required
def export_site_route(site_name):
    wb = build_site_export_workbook(site_name)
    safe_name = "".join(c for c in site_name if c.isalnum() or c in " -_")
    tmp_path = f"/tmp/{safe_name}_Report.xlsx"
    wb.save(tmp_path)
    return send_file(tmp_path, as_attachment=True, download_name=f"{safe_name}_Report.xlsx")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
